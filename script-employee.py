import pandas as pd
import os
from collections import defaultdict
import re
import openpyxl
import warnings
import datetime

# Suppress openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

data_folder = './data'  # Use relative path to the data folder
log_file_path = os.path.join(os.path.dirname(data_folder), "employee_script_log.txt")

# Setup logging
log_file = open(log_file_path, "a")
log_file.write(f"\n\n--- Script execution started at {datetime.datetime.now()} ---\n")

# Function to print to console and log file
def log_print(message):
    print(message)
    log_file.write(f"{message}\n")
    log_file.flush()  # Ensure message is written immediately

# Ensure the path is correct and accessible
if not os.path.exists(data_folder):
    log_print(f"The specified data folder does not exist: {data_folder}")
    log_file.close()
    raise FileNotFoundError(f"The specified data folder does not exist: {data_folder}")

# Dictionary to store employee salary data by year and week
# {year: {week: {employee_name: total_salary}}}
employee_salary_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
missing_dates = defaultdict(list)  # {year: [missing_dates]}

def read_excel_with_openpyxl(file_path, sheet_name="Input"):
    """Read Excel file directly with openpyxl to better handle formulas"""
    try:
        # Try using pandas with openpyxl engine first (simpler approach)
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            log_print(f"Successfully read Excel file with pandas+openpyxl, shape: {df.shape}")
            return df
        except Exception as e:
            log_print(f"Error with pandas+openpyxl: {e}")
        
        # If that fails, try direct openpyxl approach with pivot table handling
        # Load the workbook with data_only=True to get calculated values
        # and with keep_links=False to avoid pivot table issues
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
        log_print(f"Available sheets: {wb.sheetnames}")
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            log_print(f"Sheet '{sheet_name}' not found, using first sheet: {wb.sheetnames[0]}")
            ws = wb.active
        
        # Convert worksheet to dataframe
        data = []
        for row in ws.rows:
            data.append([cell.value for cell in row])
        
        # Create dataframe
        df = pd.DataFrame(data)
        
        # If first row contains headers, use them
        if len(df) > 0:
            # Use first row as header
            df.columns = df.iloc[0]
            df = df.iloc[1:].reset_index(drop=True)
        
        log_print(f"Successfully read Excel file with direct openpyxl, shape: {df.shape}")
        return df
    
    except Exception as e:
        log_print(f"Error reading with openpyxl: {e}")
        # Fall back to pandas with xlrd
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
            log_print(f"Fell back to pandas with xlrd, shape: {df.shape}")
            return df
        except Exception as e2:
            # Last resort - try with default engine
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                log_print(f"Fell back to pandas default, shape: {df.shape}")
                return df
            except Exception as e3:
                log_print(f"All reading methods failed: {e3}")
                raise

def extract_week(filename):
    # Match "APR 2-8", "APR. 2-8", "Apr 2-8", etc.
    month_date_match = re.search(r'([A-Za-z]+)\.?\s*(\d{1,2})[-–](\d{1,2})', filename)
    if month_date_match:
        month = month_date_match.group(1)
        start = month_date_match.group(2)
        end = month_date_match.group(3)
        return f"{month.upper()} {start}-{end}"
    # Fallbacks as before
    week_match = re.search(r'(?:week|w)[\s\-_]?(\d{1,2})', filename, re.IGNORECASE)
    if week_match:
        return f"Week {week_match.group(1)}"
    date_match = re.search(r'(\d{1,2})[-–](\d{1,2})', filename)
    if date_match:
        return f"{date_match.group(1)}-{date_match.group(2)}"
    return os.path.splitext(filename)[0]

def get_week_label(file_path, spreadsheet, sheet_name="Input"):
    """
    Try to get the week label from cell C1 of the Input sheet.
    If not found, fall back to extracting from the filename (ignoring the first number).
    Example filename: '14 APR. 2-8, 2025.xlsm' -> 'APR 2-8'
    """
    # Try to get from C1
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        week_label = ws.cell(row=1, column=3).value  # C1
        if week_label and str(week_label).strip():
            return str(week_label).strip()
    except Exception as e:
        log_print(f"Error reading C1 for week label in {spreadsheet}: {e}")

    # Fallback: extract from filename, ignoring the first number
    # Example: '14 APR. 2-8, 2025.xlsm'
    # Remove leading number and space
    filename = os.path.splitext(spreadsheet)[0]
    match = re.match(r'^\d+\s+(.+)$', filename)
    if match:
        filename = match.group(1)
    # Now extract month and date range
    month_date_match = re.search(r'([A-Za-z]+)\.?\s*(\d{1,2})[-–](\d{1,2})', filename)
    if month_date_match:
        month = month_date_match.group(1)
        start = month_date_match.group(2)
        end = month_date_match.group(3)
        return f"{month.upper()} {start}-{end}"
    # Fallbacks as before
    week_match = re.search(r'(?:week|w)[\s\-_]?(\d{1,2})', filename, re.IGNORECASE)
    if week_match:
        return f"Week {week_match.group(1)}"
    date_match = re.search(r'(\d{1,2})[-–](\d{1,2})', filename)
    if date_match:
        return f"{date_match.group(1)}-{date_match.group(2)}"
    return filename


# Get all spreadsheets in the data folder
spreadsheets = [file for file in os.listdir(data_folder) if file.endswith('.xlsx') or file.endswith('.xlsm')]
log_print(f"Found {len(spreadsheets)} Excel files in data folder")

for spreadsheet in spreadsheets:
    file_path = os.path.join(data_folder, spreadsheet)
    log_print(f"\n--- Processing {spreadsheet} ---")
    
    try:
        # Extract year
        match = re.search(r'(\d{4})\.xlsx[m]?$', spreadsheet)
        if match:
            year = int(match.group(1))
        else:
            match = re.findall(r'\d{4}', spreadsheet)
            if match:
                year = int(match[-1])
            else:
                log_print(f"Could not extract year from filename: {spreadsheet}. Skipping file.")
                continue

        # Use new week label function
        week = get_week_label(file_path, spreadsheet, "Input")
        log_print(f"Extracted year {year}, week {week} from C1 or filename: {spreadsheet}")

        
        # Read the sheet using enhanced openpyxl function
        try:
            df_full = read_excel_with_openpyxl(file_path, "Input")
            
            # Create debug directory
            debug_dir = os.path.join(data_folder, "../debug")
            os.makedirs(debug_dir, exist_ok=True)
            debug_file = os.path.join(debug_dir, f"openpyxl_{spreadsheet}")
            df_full.to_excel(debug_file, sheet_name="FullSheet", index=False)
            log_print(f"Saved full sheet to {debug_file}")
            
            # Find where the actual data begins
            data_start_row = None
            for i in range(len(df_full)):
                cell_value = df_full.iloc[i, 0]
                if pd.notna(cell_value) and isinstance(cell_value, str):
                    if not cell_value.lower().strip() in ['name', 'employee name', 'employee']:
                        data_start_row = i
                        log_print(f"Found data starting at row {i+1} with value: {cell_value}")
                        break
            
            if data_start_row is None:
                log_print(f"Could not find data start row. Using default of row 5.")
                data_start_row = 4
            
            # Create a dataframe with just the data
            df = df_full.iloc[data_start_row:].copy().reset_index(drop=True)
            log_print(f"Extracted data frame with shape: {df.shape}")
            
            # Try to locate salary column
            salary_column = None
            valid_columns = []
            
            # First try column AT (index 45) if it exists
            if df.shape[1] > 45:
                valid_columns.append(45)
            
            # Then look for columns with numeric values
            for col_idx in range(df.shape[1]):
                col_data = df.iloc[:, col_idx]
                numeric_count = 0
                
                for val in col_data:
                    if pd.notna(val):
                        try:
                            if isinstance(val, str):
                                # Remove currency symbols, commas, etc.
                                clean_val = re.sub(r'[^\d.-]', '', val)
                                if clean_val:
                                    float(clean_val)
                                    numeric_count += 1
                            else:
                                float(val)
                                numeric_count += 1
                        except:
                            pass
                
                if numeric_count > 2:  # At least a few valid numeric entries
                    valid_columns.append(col_idx)
            
            # Try each potential salary column
            for col_idx in valid_columns:
                employee_names = df.iloc[:, 0]  # Column A
                net_salaries = df.iloc[:, col_idx]
                valid_rows = 0
                
                for emp_name, salary in zip(employee_names, net_salaries):
                    if pd.notna(emp_name) and pd.notna(salary):
                        try:
                            if isinstance(salary, str):
                                clean_salary = re.sub(r'[^\d.-]', '', salary)
                                if clean_salary:
                                    float(clean_salary)
                                    valid_rows += 1
                            else:
                                float(salary)
                                valid_rows += 1
                        except:
                            pass
                
                if valid_rows > 0:
                    salary_column = col_idx
                    log_print(f"Using column {col_idx} with {valid_rows} valid salary entries")
                    break
            
            if salary_column is None:
                log_print(f"No valid salary column found. Skipping file.")
                continue
            
            # Process the data
            rows_processed = 0
            for i, (employee_name, net_salary) in enumerate(zip(employee_names, net_salaries)):
                if pd.notna(employee_name) and pd.notna(net_salary):
                    try:
                        if isinstance(net_salary, str):
                            clean_salary = re.sub(r'[^\d.-]', '', net_salary)
                            net_salary_float = float(clean_salary) if clean_salary else 0
                        else:
                            net_salary_float = float(net_salary)
                            
                        # Store data by year, week, and employee
                        employee_salary_data[year][week][employee_name] += net_salary_float
                        rows_processed += 1
                    except (ValueError, TypeError) as e:
                        log_print(f"Skipping invalid salary: {e}")
            
            log_print(f"Processed {rows_processed} employee records")
            
        except Exception as e:
            log_print(f"Error processing file: {e}")
            continue
        
    except Exception as e:
        log_print(f"Error processing {spreadsheet}: {e}")

# Create output file with weekly breakdown
if employee_salary_data:
    output_file = os.path.join(data_folder, "../EmployeeSummary.xlsx")
    with pd.ExcelWriter(output_file) as writer:
        # Process each year
        for year, weeks_data in employee_salary_data.items():
            log_print(f"Creating summary for year {year} with {len(weeks_data)} weeks")
            
            # Get all employee names across all weeks
            all_employees = set()
            for week_data in weeks_data.values():
                all_employees.update(week_data.keys())
            
            # Get sorted list of weeks
            weeks_sorted = sorted(weeks_data.keys())
            
            # Create a DataFrame with columns for each week plus a total
            data = []
            for employee in sorted(all_employees):
                row = {'Employee Name': employee}
                total_salary = 0
                
                # Add data for each week
                for week in weeks_sorted:
                    week_label = f"Week {week}" if isinstance(week, int) else str(week)
                    salary = weeks_data[week][employee]
                    row[week_label] = salary
                    total_salary += salary
                
                # Add total
                row['Total'] = total_salary
                data.append(row)
            
            # Create and save DataFrame
            columns = ['Employee Name'] + [f"Week {w}" if isinstance(w, int) else str(w) for w in weeks_sorted] + ['Total']
            df = pd.DataFrame(data, columns=columns)
            df.to_excel(writer, index=False, sheet_name=str(year))
            
            log_print(f"Created summary for {year} with {len(df)} employees and {len(weeks_sorted)} weeks")
    
    log_print(f"Summary saved successfully to {output_file}")
else:
    log_print("No data was processed. Check your file format and content.")

# Close the log file
log_print(f"\n--- Script execution completed at {datetime.datetime.now()} ---")
log_file.close()