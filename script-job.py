import pandas as pd
import os
import datetime
import openpyxl
import re

data_folder = './data'  # Use relative path to the data folder
log_file_path = os.path.join(os.path.dirname(data_folder), "job_script_log.txt")

# Setup logging
log_file = open(log_file_path, "a")
log_file.write(f"\n\n--- Script execution started at {datetime.datetime.now()} ---\n")

# Function to print to console and log file
def log_print(message):
    print(message)
    log_file.write(f"{message}\n")
    log_file.flush()  # Ensure message is written immediately

# Helper to get week label from C1 or fallback to filename
def get_week_label(file_path, spreadsheet, sheet_name="Bossing"):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, keep_links=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        week_label = ws.cell(row=1, column=3).value  # C1
        if week_label and str(week_label).strip():
            return str(week_label).strip()
    except Exception as e:
        log_print(f"Error reading C1 for week label in {spreadsheet}: {e}")

    # Fallback: extract from filename, ignoring the first number
    filename = os.path.splitext(spreadsheet)[0]
    match = re.match(r'^\d+\s+(.+)$', filename)
    if match:
        filename = match.group(1)
    month_date_match = re.search(r'([A-Za-z]+)\.?\s*(\d{1,2})[-–](\d{1,2})', filename)
    if month_date_match:
        month = month_date_match.group(1)
        start = month_date_match.group(2)
        end = month_date_match.group(3)
        return f"{month.upper()} {start}-{end}"
    week_match = re.search(r'(?:week|w)[\s\-_]?(\d{1,2})', filename, re.IGNORECASE)
    if week_match:
        return f"Week {week_match.group(1)}"
    date_match = re.search(r'(\d{1,2})[-–](\d{1,2})', filename)
    if date_match:
        return f"{date_match.group(1)}-{date_match.group(2)}"
    return filename

# Ensure the path is correct and accessible
if not os.path.exists(data_folder):
    log_print(f"The specified data folder does not exist: {data_folder}")
    log_file.close()
    raise FileNotFoundError(f"The specified data folder does not exist: {data_folder}")

# Dictionary to store the extracted data per week
# job_order_data[job_order][week_label] = expense
job_order_data = {}
all_weeks = set()

# Get all spreadsheets in the data folder
spreadsheets = [file for file in os.listdir(data_folder) if file.endswith('.xlsx') or file.endswith('.xlsm')]
log_print(f"Found {len(spreadsheets)} Excel files in data folder")

for spreadsheet in spreadsheets:
    file_path = os.path.join(data_folder, spreadsheet)
    log_print(f"\nProcessing {spreadsheet}...")

    week_label = get_week_label(file_path, spreadsheet, "Bossing")
    all_weeks.add(week_label)
    log_print(f"Using week label: {week_label}")

    # Read the "Bossing" sheet starting from row 5
    try:
        df = pd.read_excel(file_path, sheet_name="Bossing", skiprows=3)
        job_orders = df.iloc[:, 0]  # Column A: Job Order Name
        total_expenses = df.iloc[:, 1]  # Column B: Total Expense for Employees

        processed_count = 0
        for job_order, total_expense in zip(job_orders, total_expenses):
            if pd.notna(job_order) and pd.notna(total_expense):
                if str(job_order).strip().lower() == "grand total":
                    continue
                if job_order not in job_order_data:
                    job_order_data[job_order] = {}
                job_order_data[job_order][week_label] = total_expense
                processed_count += 1

        log_print(f"Processed {spreadsheet} successfully - found {processed_count} valid job order entries.")
    except Exception as e:
        log_print(f"Error processing {spreadsheet}: {e}")

# Create a DataFrame from the job_order_data dictionary with columns per week and a total
if job_order_data:
    weeks_sorted = sorted(all_weeks)
    summary_data = []
    for job_order, week_expenses in job_order_data.items():
        row = {'Job Order': job_order}
        total = 0
        for week in weeks_sorted:
            expense = week_expenses.get(week, 0)
            row[week] = expense
            total += expense
        row['Total'] = total
        summary_data.append(row)
    columns = ['Job Order'] + list(weeks_sorted) + ['Total']
    summary_df = pd.DataFrame(summary_data, columns=columns)
    summary_df = summary_df.sort_values('Job Order')
    log_print(f"\nCreated summary dataframe with {len(summary_df)} job orders and {len(weeks_sorted)} weeks")

    output_file = os.path.join(data_folder, "../Summary.xlsx")
    try:
        summary_df.to_excel(output_file, index=False, sheet_name="Summary")
        log_print(f"Summary saved successfully to {output_file}")
    except Exception as e:
        log_print(f"Error saving summary: {e}")
else:
    log_print("No job order data found.")

log_print(f"\n--- Script execution completed at {datetime.datetime.now()} ---")
log_file.close()